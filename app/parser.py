"""
XLSX billing data parser — general-purpose column detection and normalization.
"""
import re
import json
from pathlib import Path
from typing import Any

import openpyxl


# Column detection heuristics: (canonical_name, [keyword_aliases])
COLUMN_PATTERNS = [
    ("date",              ["时间", "日期", "date", "time", "month"]),
    ("resource_name",     ["资源名称", "name", "resource_name"]),
    ("resource_id",       ["资源id", "resourceid", "resource_id", "apikey", "api_key", "key", "token"]),
    ("billing_method",    ["计费方式", "billing_method"]),
    ("resource_type",     ["资源类型", "type", "resource_type"]),
    ("model",             ["模型", "model"]),
    ("usage_desc",        ["配置描述", "用量描述", "用量", "usage", "消耗", "描述"]),
    ("site",              ["站点", "site", "region", "区域", "地域"]),
    ("transaction_type",  ["交易类型", "transaction_type"]),
    ("service_fee",       ["服务费", "service_fee"]),
    ("cost",              ["费用", "cost", "金额", "price"]),
]

# Token type patterns in usage descriptions (order matters: specific first)
TOKEN_TYPE_PATTERNS = [
    (re.compile(r"(缓存输入|缓存命中|cache[_\s]?hit|cached)[:\s：]*([\d,]+)\s*tokens?", re.IGNORECASE), "cache_hit"),
    (re.compile(r"(输入|input)[:\s：]*([\d,]+)\s*tokens?", re.IGNORECASE), "input"),
    (re.compile(r"(输出|output)[:\s：]*([\d,]+)\s*tokens?", re.IGNORECASE), "output"),
]

# --- Token-stream parser patterns (for plain-text billing dumps) ---
# Each whitespace-separated token is classified into one of these categories.
# Headers and trailers are naturally ignored: their tokens don't match any
# structured pattern and fall through as unrecognized general strings.

# Token-type prefix + count, e.g. "输入：308,574tokens" or "输出:100tokens"
_TEXT_TOKEN_COUNT_RE = re.compile(
    r'^(缓存输入|输入|输出)[：:]\s*([\d,]+)\s*tokens$'
)
# Tag + optional value, e.g. "名称：harry-opencode" or "模型：DeepSeek-V4-Pro"
_TEXT_TAG_RE = re.compile(r'^(.+?)[：:](.*)$')
# Date: "2026-07-16"
_TEXT_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
# Time: "19:00:00"
_TEXT_TIME_RE = re.compile(r'^\d{2}:\d{2}:\d{2}$')
# Cost: decimal number like "3.70289" (must have decimal point to distinguish
# from integer token-counts or other bare numbers)
_TEXT_COST_RE = re.compile(r'^\d+\.\d+$')

# Maps the Chinese token-type prefix to the canonical token type name
_TEXT_TOKEN_TYPE_MAP = {
    '缓存输入': 'cache_hit',
    '输入': 'input',
    '输出': 'output',
}

# Reverse map for building usage_desc strings
_TEXT_TYPE_PREFIX_MAP = {v: k for k, v in _TEXT_TOKEN_TYPE_MAP.items()}


def _parse_float(val) -> float:
    """Parse a value to float, handling strings with commas."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def estimate_model_prices(records: list[dict]) -> dict[str, dict[str, float | None]]:
    """Estimate per-token prices for each model using least-squares linear regression.

    For each model, solves the overdetermined system:
        input_i * p_in + output_i * p_out + cache_i * p_cache = cost_i
    using numpy least-squares (minimizing ||Ax - b||^2).
    Requires at least 2 rows per model. Falls back to simple averaging for single-type rows
    if the regression produces implausible (negative or zero) results.
    """
    import numpy as np

    models = set(r["model"] for r in records if r["model"])
    prices: dict[str, dict[str, float | None]] = {}

    for model in models:
        model_recs = [r for r in records if r["model"] == model]
        prices[model] = {"input": None, "output": None, "cache_hit": None}

        # Build matrix A (n×3) and vector b (n×1)
        rows_A = []
        rows_b = []
        for r in model_recs:
            if r["tokens_total"] == 0 and r["cost"] == 0:
                continue
            rows_A.append([float(r["tokens_input"]), float(r["tokens_output"]), float(r["tokens_cache_hit"])])
            rows_b.append(float(r["cost"]))

        if len(rows_A) < 2:
            continue

        A = np.array(rows_A, dtype=np.float64)
        b = np.array(rows_b, dtype=np.float64)

        try:
            x, residuals, rank, sv = np.linalg.lstsq(A, b, rcond=None)
        except np.linalg.LinAlgError:
            continue

        # Only accept positive price estimates
        if x[0] > 0:
            prices[model]["input"] = float(x[0])
        if x[1] > 0:
            prices[model]["output"] = float(x[1])
        if x[2] > 0:
            prices[model]["cache_hit"] = float(x[2])

        # Fallback: if a price type couldn't be estimated via regression but
        # there are single-type rows, use the simple average.
        for token_type, col_idx in [("input", 0), ("output", 1), ("cache_hit", 2)]:
            if prices[model][token_type] is not None:
                continue
            token_key = f"tokens_{token_type}"
            single_recs = [
                r for r in model_recs
                if r[token_key] > 0
                and all(r[f"tokens_{other}"] == 0 for other in ["input", "output", "cache_hit"] if other != token_type)
            ]
            if single_recs:
                total_tokens = sum(r[token_key] for r in single_recs)
                total_cost = sum(r["cost"] for r in single_recs)
                if total_tokens > 0:
                    prices[model][token_type] = total_cost / total_tokens

    return prices


def detect_columns(headers: list[str]) -> dict[str, int]:
    """Map canonical column names to 0-based indices using heuristics."""
    mapping: dict[str, int] = {}
    normalized_headers = [h.strip().lower() if h else "" for h in headers]

    for canonical, aliases in COLUMN_PATTERNS:
        for idx, h in enumerate(normalized_headers):
            if canonical in mapping:
                break
            for alias in aliases:
                if alias in h:
                    mapping[canonical] = idx
                    break

    return mapping


def parse_usage_desc(text: str) -> list[dict[str, Any]]:
    """Parse a usage description cell into token type + count entries.
    
    Returns at most one entry — the first matching pattern (ordered by specificity).
    """
    if not text or not isinstance(text, str):
        return []

    for pattern, token_type in TOKEN_TYPE_PATTERNS:
        m = pattern.search(text)
        if m:
            count_str = m.group(2).replace(",", "")
            try:
                count = int(count_str)
            except ValueError:
                continue
            return [{"type": token_type, "tokens": count}]

    return []


def _ensure_xlsx(filepath: Path) -> Path:
    """If the file is an xlsx with a non-xlsx extension, copy to a temp .xlsx file."""
    if filepath.suffix.lower() in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return filepath
    # Check magic bytes — xlsx files are ZIP archives starting with PK
    with open(filepath, 'rb') as f:
        header = f.read(4)
    if header[:2] != b'PK':
        raise ValueError(f"File is not a valid xlsx: {filepath.name}")
    # Copy to temp .xlsx
    import tempfile, shutil
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    shutil.copyfile(filepath, tmp.name)
    return Path(tmp.name)


def parse_xlsx(filepath: str | Path) -> dict[str, Any]:
    """Parse an xlsx billing file and return normalized JSON-serializable data."""
    filepath = Path(filepath)
    actual_path = _ensure_xlsx(filepath)
    try:
        wb = openpyxl.load_workbook(actual_path)
    except Exception:
        if actual_path != filepath:
            actual_path.unlink(missing_ok=True)
        raise
    try:
        ws = wb.active

        raw_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))

        def cell_val(cell):
            return cell.value

        rows = [[cell_val(c) for c in row] for row in raw_rows]

        if not rows:
            wb.close()
            return {"error": "Empty spreadsheet", "records": []}

        # Detect headers from first row
        headers = [str(c) if c is not None else "" for c in rows[0]]
        col_map = detect_columns(headers)

        records = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            values = [str(c) if c is not None else "" for c in row]

            usage_desc = values[col_map["usage_desc"]] if "usage_desc" in col_map else ""
            token_entries = parse_usage_desc(usage_desc)

            record = {
                "date": values[col_map["date"]] if "date" in col_map else "",
                "resource_name": values[col_map["resource_name"]] if "resource_name" in col_map else "",
                "resource_type": values[col_map["resource_type"]] if "resource_type" in col_map else "",
                "resource_id": values[col_map["resource_id"]] if "resource_id" in col_map else "",
                "billing_method": values[col_map["billing_method"]] if "billing_method" in col_map else "",
                "model": values[col_map["model"]] if "model" in col_map else "",
                "usage_desc": usage_desc,
                "site": values[col_map["site"]] if "site" in col_map else "",
                "transaction_type": values[col_map["transaction_type"]] if "transaction_type" in col_map else "",
                "service_fee": _parse_float(values[col_map["service_fee"]]) if "service_fee" in col_map else 0.0,
                "cost": _parse_float(values[col_map["cost"]]) if "cost" in col_map else 0.0,
                "tokens": token_entries,
            }
            record["tokens_input"] = sum(e["tokens"] for e in token_entries if e["type"] == "input")
            record["tokens_output"] = sum(e["tokens"] for e in token_entries if e["type"] == "output")
            record["tokens_cache_hit"] = sum(e["tokens"] for e in token_entries if e["type"] == "cache_hit")
            record["tokens_total"] = record["tokens_input"] + record["tokens_output"] + record["tokens_cache_hit"]
            records.append(record)

        wb.close()

        # Build aggregations via shared function
        return aggregate_records(records, meta={
            "filename": filepath.name,
            "column_map": {k: headers[v] for k, v in col_map.items()},
        })
    finally:
        if actual_path != filepath:
            actual_path.unlink(missing_ok=True)


def aggregate_records(records: list[dict], meta: dict | None = None) -> dict[str, Any]:
    """Aggregate a list of record dicts into summary, by_key, by_resource_name, by_model, timeline.

    This is the same logic used inside parse_xlsx, extracted for reuse when merging
    records from multiple files.

    Args:
        records: List of record dicts (as produced by parse_xlsx per-row logic).
        meta: Optional extra metadata to include in the top-level result.

    Returns:
        A dict with keys: meta, summary, records, by_key, by_resource_name, by_model,
        timeline, prices.
    """
    api_keys = sorted(set(r["resource_id"] for r in records if r["resource_id"]))
    models = sorted(set(r["model"] for r in records if r["model"]))
    dates = sorted(set(r["date"] for r in records if r["date"]))
    resource_names = sorted(set(r["resource_name"] for r in records if r["resource_name"]))

    # By API key
    by_key = {}
    for key in api_keys:
        kr = [r for r in records if r["resource_id"] == key]
        by_key[key] = {
            "resource_type": kr[0]["resource_type"] if kr else "",
            "record_count": len(kr),
            "tokens_input": sum(r["tokens_input"] for r in kr),
            "tokens_output": sum(r["tokens_output"] for r in kr),
            "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in kr),
            "tokens_total": sum(r["tokens_total"] for r in kr),
            "cost": sum(r["cost"] for r in kr),
            "models": sorted(set(r["model"] for r in kr)),
        }

    # By resource_name (user-friendly labels)
    by_resource_name = {}
    for name in resource_names:
        nr = [r for r in records if r["resource_name"] == name]
        by_resource_name[name] = {
            "record_count": len(nr),
            "tokens_input": sum(r["tokens_input"] for r in nr),
            "tokens_output": sum(r["tokens_output"] for r in nr),
            "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in nr),
            "tokens_total": sum(r["tokens_total"] for r in nr),
            "cost": sum(r["cost"] for r in nr),
            "models": sorted(set(r["model"] for r in nr)),
            "api_keys": sorted(set(r["resource_id"] for r in nr)),
        }

    # By model
    by_model = {}
    for model in models:
        mr = [r for r in records if r["model"] == model]
        by_model[model] = {
            "record_count": len(mr),
            "tokens_input": sum(r["tokens_input"] for r in mr),
            "tokens_output": sum(r["tokens_output"] for r in mr),
            "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in mr),
            "tokens_total": sum(r["tokens_total"] for r in mr),
            "cost": sum(r["cost"] for r in mr),
        }

    # Timeline by date
    timeline = {}
    for date in dates:
        dr = [r for r in records if r["date"] == date]
        entry = {
            "tokens_input": sum(r["tokens_input"] for r in dr),
            "tokens_output": sum(r["tokens_output"] for r in dr),
            "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in dr),
            "tokens_total": sum(r["tokens_total"] for r in dr),
            "cost": sum(r["cost"] for r in dr),
            "by_key": {},
        }
        for key in api_keys:
            kdr = [r for r in dr if r["resource_id"] == key]
            if kdr:
                entry["by_key"][key] = {
                    "tokens_input": sum(r["tokens_input"] for r in kdr),
                    "tokens_output": sum(r["tokens_output"] for r in kdr),
                    "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in kdr),
                    "tokens_total": sum(r["tokens_total"] for r in kdr),
                    "cost": sum(r["cost"] for r in kdr),
                }
        timeline[date] = entry

    summary = {
        "total_records": len(records),
        "api_key_count": len(api_keys),
        "model_count": len(models),
        "date_count": len(dates),
        "tokens_input": sum(r["tokens_input"] for r in records),
        "tokens_output": sum(r["tokens_output"] for r in records),
        "tokens_cache_hit": sum(r["tokens_cache_hit"] for r in records),
        "tokens_total": sum(r["tokens_total"] for r in records),
        "cost": sum(r["cost"] for r in records),
    }

    result = {
        "meta": meta or {},
        "summary": summary,
        "records": records,
        "by_key": by_key,
        "by_resource_name": by_resource_name,
        "by_model": by_model,
        "timeline": timeline,
        "prices": estimate_model_prices(records),
    }
    # Always include these in meta for downstream consumers
    if meta:
        result["meta"].update({
            "api_keys": api_keys,
            "resource_names": resource_names,
            "models": models,
            "dates": dates,
        })
    else:
        result["meta"] = {
            "api_keys": api_keys,
            "resource_names": resource_names,
            "models": models,
            "dates": dates,
        }

    return result


def parse_text(filepath: str | Path) -> dict[str, Any]:
    """Parse a plain-text billing dump using a robust token-stream approach.

    The file is treated as a flat stream of whitespace-separated tokens. Each
    token is classified in priority order:

      1. Token-type + count  (e.g. "输入：308,574tokens"  or  "缓存输入:100tokens")
      2. Tag + value          (e.g. "名称：harry-opencode"  or  "模型：GLM-5.2")
      3. Date                 (e.g. "2026-07-16")
      4. Time                 (e.g. "19:00:00"  — combines with preceding date)
      5. Cost                 (decimal number, e.g. "3.70289")
      6. General string        (everything else — headers, trailers, noise — ignored)

    Section headers ("充值优惠", "资源ID账单", "明细账单", …), column headers
    ("资源名称", "配置描述", …), and footers ("•••", "欢迎咨询") are never
    matched explicitly — their tokens simply don't fit any structured pattern
    and fall through as unrecognized noise.

    A record is emitted when all six fields are present:
    resource_name, model, token_type, token_count, timestamp, cost.
    The "名称：" tag signals the start of a new record group and resets the
    buffer, so malformed/incomplete groups are discarded rather than corrupting
    the next record.  Duplicate records are deduplicated by full signature.
    """
    filepath = Path(filepath)

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Flatten to a stream of whitespace-separated tokens (spaces, tabs, newlines)
    tokens = content.split()

    # State machine buffer for the current record
    cur = {
        'resource_name': None,
        'model': None,
        'token_type': None,
        'token_count': None,
        'timestamp': None,
        'cost': None,
    }
    pending_date = None      # date waiting for a time token to complete it
    pending_tag = None       # tag whose value will arrive as the next general string

    records = []
    seen = set()             # dedup signatures

    def reset():
        """Discard the current buffer and start fresh."""
        nonlocal pending_date, pending_tag
        for k in cur:
            cur[k] = None
        pending_date = None
        pending_tag = None

    def try_emit():
        """Emit the current record if all six fields are present, then reset."""
        if all(v is not None for v in cur.values()):
            sig = (
                cur['resource_name'],
                cur['model'],
                cur['token_type'],
                cur['token_count'],
                cur['timestamp'],
                cur['cost'],
            )
            if sig not in seen:
                seen.add(sig)
                token_count = cur['token_count']
                token_type = cur['token_type']
                type_prefix = _TEXT_TYPE_PREFIX_MAP.get(token_type, '')
                usage_desc = f"{type_prefix}：{format(token_count, ',')}tokens"
                date = cur['timestamp'][:10]
                records.append({
                    "date": date,
                    "resource_name": cur['resource_name'],
                    "resource_type": "",
                    "resource_id": "",
                    "billing_method": "",
                    "model": cur['model'],
                    "usage_desc": usage_desc,
                    "site": "",
                    "transaction_type": "",
                    "service_fee": 0.0,
                    "cost": cur['cost'],
                    "tokens": [{"type": token_type, "tokens": token_count}],
                    "tokens_input": token_count if token_type == "input" else 0,
                    "tokens_output": token_count if token_type == "output" else 0,
                    "tokens_cache_hit": token_count if token_type == "cache_hit" else 0,
                    "tokens_total": token_count,
                })
            reset()

    for tok in tokens:
        # 1. Token-type + count (highest priority — has both type prefix and "tokens" suffix)
        m = _TEXT_TOKEN_COUNT_RE.match(tok)
        if m:
            type_prefix = m.group(1)
            token_type = _TEXT_TOKEN_TYPE_MAP.get(type_prefix)
            if token_type:
                try:
                    token_count = int(m.group(2).replace(',', ''))
                    cur['token_type'] = token_type
                    cur['token_count'] = token_count
                except ValueError:
                    pass
            pending_tag = None
            continue

        # 2. Tag + value  (e.g. "名称：harry-opencode", "模型：GLM-5.2")
        m = _TEXT_TAG_RE.match(tok)
        if m:
            tag, value = m.group(1), m.group(2)
            if tag == '名称':
                # Start of a new record group — discard any incomplete state
                reset()
                if value:
                    cur['resource_name'] = value
                else:
                    pending_tag = '名称'
            elif tag == '模型':
                if value:
                    cur['model'] = value
                else:
                    pending_tag = '模型'
            # Unknown tags (时间, 总支出, etc.) are silently ignored
            continue

        # 3. Date  (e.g. "2026-07-16")
        if _TEXT_DATE_RE.match(tok):
            # If a previous date never got a time, treat it as a bare timestamp
            if pending_date and cur['timestamp'] is None:
                cur['timestamp'] = pending_date
            pending_date = tok
            pending_tag = None
            continue

        # 4. Time  (e.g. "19:00:00") — combines with pending date into a timestamp
        if _TEXT_TIME_RE.match(tok):
            if pending_date:
                cur['timestamp'] = f"{pending_date} {tok}"
                pending_date = None
            pending_tag = None
            continue

        # 5. Cost  (decimal number — must have "." to avoid matching bare integers)
        if _TEXT_COST_RE.match(tok):
            try:
                cur['cost'] = float(tok)
            except ValueError:
                continue
            # If we have a pending date but no timestamp, use the date alone
            if cur['timestamp'] is None and pending_date:
                cur['timestamp'] = pending_date
                pending_date = None
            try_emit()
            continue

        # 6. General string
        # If a tag is waiting for its value, this token is the value
        if pending_tag:
            if pending_tag == '名称':
                cur['resource_name'] = tok
            elif pending_tag == '模型':
                cur['model'] = tok
            pending_tag = None
        # Otherwise: unrecognized token (headers, trailers, noise) — ignored

    if not records:
        return {"error": "No records found in text file", "records": []}

    return aggregate_records(records, meta={
        "filename": filepath.name,
        "format": "text",
    })


def parse_billing_file(filepath: str | Path) -> dict[str, Any]:
    """Dispatcher: detect file format (xlsx vs plain text) and route to the right parser.

    Detects by PK ZIP magic bytes (xlsx) vs anything else (text). If the file
    has a known xlsx/xls extension, also treats it as xlsx.
    """
    filepath = Path(filepath)

    # Fast-path by extension
    if filepath.suffix.lower() in ('.txt',):
        return parse_text(filepath)

    # Sniff magic bytes
    try:
        with open(filepath, 'rb') as f:
            header = f.read(4)
        if header[:2] == b'PK':
            return parse_xlsx(filepath)
    except OSError:
        pass

    # Fallback: treat as text if extension is not a known xlsx variant
    if filepath.suffix.lower() in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'):
        return parse_xlsx(filepath)
    return parse_text(filepath)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python parser.py <file.xlsx>")
        sys.exit(1)

    result = parse_xlsx(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
